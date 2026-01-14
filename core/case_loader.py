"""Excel测试用例加载器"""
import openpyxl
from typing import List, Optional, Union
from dataclasses import dataclass
from pathlib import Path

from utils.logger import get_logger

logger = get_logger(__name__)


@dataclass
class TestCase:
    """测试用例数据类

    Attributes:
        case_id: 测试用例ID
        module: 功能模块
        api_name: 接口名称
        url: 请求地址（相对路径）
        pre_condition: 前置条件（数据提取规则JSON字符串）
        method: 请求方法（GET/POST/PUT/DELETE）
        param_type: 请求参数类型（params/data/json）
        params: 请求参数（JSON字符串）
        expected_result: 期望结果（JSON字符串）
        is_run: 是否运行（Y/N）
        headers: 请求头（JSON字符串）
        expected_status: 期望HTTP状态码
        performance_config: 性能配置（JSON字符串，可选）
        max_response_time: 最大响应时间（毫秒，可选）
    """
    case_id: str
    module: str
    api_name: str
    url: str
    pre_condition: str
    method: str
    param_type: str
    params: str
    expected_result: str
    is_run: str
    headers: str
    expected_status: int
    performance_config: str = "{}"  # 默认为空配置
    max_response_time: int = 0  # 默认为0表示不限制


class CaseLoader:
    """测试用例加载器

    支持加载单个或多个Excel文件
    支持加载单个sheet、多个sheet或所有sheet
    """

    def __init__(self, excel_path: str, sheet_names: Union[str, List[str]] = "all"):
        """初始化用例加载器

        Args:
            excel_path: Excel文件路径或目录路径
            sheet_names: 工作表名称，"all"表示加载所有sheet，也可传入sheet名称列表
        """
        self.excel_path = Path(excel_path)
        # 处理sheet_names参数
        if isinstance(sheet_names, str):
            self.sheet_names = [sheet_names] if sheet_names != "all" else None
        else:
            self.sheet_names = sheet_names
        self.logger = logger

    def load_cases(self) -> List[TestCase]:
        """加载所有测试用例

        Returns:
            测试用例列表
        """
        # 判断是文件还是目录
        if self.excel_path.is_file():
            # 单个文件
            return self._load_single_file(self.excel_path)
        elif self.excel_path.is_dir():
            # 目录，加载所有Excel文件
            return self._load_directory(self.excel_path)
        else:
            raise FileNotFoundError(f"路径不存在: {self.excel_path}")

    def _load_single_file(self, file_path: Path) -> List[TestCase]:
        """加载单个Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            测试用例列表
        """
        self.logger.info(f"加载Excel文件: {file_path.name}")

        if not file_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path)

            # 获取要加载的sheet列表
            if self.sheet_names is None:
                # 加载所有sheet
                sheets_to_load = workbook.sheetnames
                self.logger.info(f"加载所有sheet: {sheets_to_load}")
            else:
                # 加载指定的sheet
                sheets_to_load = [s for s in self.sheet_names if s in workbook.sheetnames]
                if not sheets_to_load:
                    raise ValueError(f"指定的sheet不存在: {self.sheet_names}，文件中的sheet: {workbook.sheetnames}")
                self.logger.info(f"加载指定的sheet: {sheets_to_load}")

            all_cases = []
            # 遍历每个sheet
            for sheet_name in sheets_to_load:
                sheet = workbook[sheet_name]
                self.logger.info(f"正在加载sheet: {sheet_name}")

                # 跳过表头，从第2行开始
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    case = self._parse_row(row, file_path.name, sheet_name)
                    if case and case.is_run.upper() == 'Y':
                        all_cases.append(case)
                        self.logger.debug(f"加载测试用例: [{case.case_id}] {case.api_name}")

            workbook.close()
            self.logger.info(f"从 {file_path.name} 的 {len(sheets_to_load)} 个sheet中加载了 {len(all_cases)} 条测试用例")

            return all_cases

        except Exception as e:
            self.logger.error(f"加载Excel文件失败 {file_path.name}: {e}")
            raise

    def _load_directory(self, dir_path: Path) -> List[TestCase]:
        """加载目录中的所有Excel文件

        Args:
            dir_path: 目录路径

        Returns:
            测试用例列表
        """
        self.logger.info(f"从目录加载Excel文件: {dir_path}")

        # 查找所有Excel文件
        excel_files = list(dir_path.glob("*.xlsx")) + list(dir_path.glob("*.xls"))

        if not excel_files:
            raise FileNotFoundError(f"目录中没有找到Excel文件: {dir_path}")

        self.logger.info(f"找到 {len(excel_files)} 个Excel文件")

        all_cases = []
        for excel_file in excel_files:
            try:
                cases = self._load_single_file(excel_file)
                all_cases.extend(cases)
            except Exception as e:
                self.logger.warning(f"跳过文件 {excel_file.name}: {e}")
                continue

        self.logger.info(f"共加载 {len(all_cases)} 条测试用例")

        return all_cases

    def _parse_row(self, row: tuple, file_name: str = "", sheet_name: str = "Sheet1") -> Optional[TestCase]:
        """解析Excel行数据

        Args:
            row: Excel行数据元组
            file_name: 文件名（用于标识用例来源）
            sheet_name: sheet名称（用于标识用例来源）

        Returns:
            TestCase对象，解析失败返回None
        """
        try:
            # 检查行是否为空
            if not row or not any(row):
                return None

            case_id = str(row[0] if row[0] else "")

            # 如果case_id为空，生成一个默认的（包含文件名和sheet名）
            if not case_id:
                case_id = f"AUTO_{file_name}_{sheet_name}_{row[0] if row[0] else 'UNKNOWN'}"

            return TestCase(
                case_id=case_id,
                module=str(row[1] if row[1] else ""),
                api_name=str(row[2] if row[2] else ""),
                url=str(row[3] if row[3] else ""),
                pre_condition=str(row[4] if row[4] is not None else ""),
                method=str(row[5] if row[5] else "GET").upper(),
                param_type=str(row[6] if row[6] else "json"),
                params=str(row[7] if row[7] is not None else "{}"),
                expected_result=str(row[8] if row[8] is not None else "{}"),
                is_run=str(row[9] if row[9] else "Y"),
                headers=str(row[10] if row[10] is not None else "{}"),
                expected_status=int(row[11] if row[11] else 200),
                performance_config=str(row[12] if len(row) > 12 and row[12] is not None else "{}"),
                max_response_time=int(row[13] if len(row) > 13 and row[13] is not None else 0)
            )
        except Exception as e:
            self.logger.warning(f"解析行数据失败: {row}, 错误: {e}")
            return None


class MultiFileCaseLoader:
    """多文件用例加载器

    支持选择性加载指定的Excel文件
    支持加载单个sheet、多个sheet或所有sheet
    """

    def __init__(self, excel_paths: Union[str, List[str]], sheet_names: Union[str, List[str]] = "all"):
        """初始化多文件用例加载器

        Args:
            excel_paths: Excel文件路径列表或目录路径
            sheet_names: 工作表名称，"all"表示加载所有sheet，也可传入sheet名称列表
        """
        self.sheet_names = sheet_names
        self.logger = logger

        # 处理输入路径
        if isinstance(excel_paths, str):
            self.excel_paths = [Path(excel_paths)]
        else:
            self.excel_paths = [Path(p) for p in excel_paths]

    def load_cases(self) -> List[TestCase]:
        """加载所有测试用例

        Returns:
            测试用例列表
        """
        all_cases = []

        for path in self.excel_paths:
            loader = CaseLoader(str(path), self.sheet_names)
            try:
                cases = loader.load_cases()
                all_cases.extend(cases)
            except Exception as e:
                self.logger.error(f"加载文件失败 {path}: {e}")
                continue

        self.logger.info(f"共加载 {len(all_cases)} 条测试用例")

        return all_cases
